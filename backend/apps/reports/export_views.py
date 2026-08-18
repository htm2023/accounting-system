from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from apps.parties.models import Party
from apps.fiscal.models import FiscalPeriod
from .views import (
    get_trial_balance_data,
    get_income_statement_data,
    get_balance_sheet_data,
    get_cash_flow_data,
    get_party_statement_data,
)


def export_to_excel(ws, headers, rows, title):
    """Populate worksheet with headers and rows, with basic styling."""
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    ws.append([])  # blank row

    ws.append(headers)
    header_row_num = ws.max_row
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row in rows:
        ws.append(row)

    # Auto width columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


class TrialBalanceExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_period_id = request.query_params.get('fiscal_period')
        try:
            data = get_trial_balance_data(fiscal_period_id)
        except FiscalPeriod.DoesNotExist:
            return Response({'error': 'Fiscal period not found.'}, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = "ميزان المراجعة"
        headers = ['كود الحساب', 'اسم الحساب', 'النوع', 'مدين', 'دائن', 'الرصيد']
        rows = [
            [d['account_code'], d['account_name_ar'], d['account_type'], float(d['total_debit']), float(d['total_credit']), float(d['balance'])]
            for d in data
        ]
        export_to_excel(ws, headers, rows, "ميزان المراجعة")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=trial_balance.xlsx'
        wb.save(response)
        return response


class IncomeStatementExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_period_id = request.query_params.get('fiscal_period')
        data = get_income_statement_data(fiscal_period_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "قائمة الدخل"
        headers = ['البند', 'القيمة']
        rows = [
            ['الإيرادات', float(data['revenue'])],
            ['المصروفات', float(data['expenses'])],
            ['صافي الربح', float(data['net_profit'])],
        ]
        export_to_excel(ws, headers, rows, "قائمة الدخل")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=income_statement.xlsx'
        wb.save(response)
        return response


class BalanceSheetExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_period_id = request.query_params.get('fiscal_period')
        try:
            data = get_balance_sheet_data(fiscal_period_id)
        except FiscalPeriod.DoesNotExist:
            return Response({'error': 'Fiscal period not found.'}, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = "الميزانية العمومية"
        headers = ['البند', 'القيمة']
        rows = [
            ['الأصول', float(data['assets'])],
            ['الالتزامات', float(data['liabilities'])],
            ['حقوق الملكية', float(data['equity'])],
            ['إجمالي الالتزامات وحقوق الملكية', float(data['total_liabilities_equity'])],
        ]
        export_to_excel(ws, headers, rows, "الميزانية العمومية")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=balance_sheet.xlsx'
        wb.save(response)
        return response


class CashFlowExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fiscal_period_id = request.query_params.get('fiscal_period')
        data = get_cash_flow_data(fiscal_period_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "التدفقات النقدية"
        headers = ['البند', 'القيمة']
        rows = [
            ['التدفقات الداخلة', float(data['cash_inflow'])],
            ['التدفقات الخارجة', float(data['cash_outflow'])],
            ['صافي التدفق النقدي', float(data['net_cash_flow'])],
        ]
        export_to_excel(ws, headers, rows, "التدفقات النقدية")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=cash_flow.xlsx'
        wb.save(response)
        return response


class PartyStatementExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, party_id):
        try:
            data = get_party_statement_data(party_id)
        except Party.DoesNotExist:
            return Response({'error': 'Party not found.'}, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = "كشف حساب"
        headers = ['التاريخ', 'الوصف', 'مدين', 'دائن', 'الرصيد']
        rows = [
            [str(e['date']), e['description'], float(e['debit']), float(e['credit']), float(e['balance'])]
            for e in data['entries']
        ]
        export_to_excel(ws, headers, rows, f"كشف حساب: {data['party']}")
        # Add opening and closing balance rows
        ws.append([])
        ws.append(['الرصيد الافتتاحي', '', '', '', float(data['opening_balance'])])
        ws.append(['الرصيد الختامي', '', '', '', float(data['closing_balance'])])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=party_statement_{party_id}.xlsx'
        wb.save(response)
        return response
